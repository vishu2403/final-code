"""Student management routes for roster uploads and profile syncing."""
from __future__ import annotations

import base64
import csv
import io
import re
from typing import Any, Dict, Iterable, List, Optional
from typing import Any, Dict, Iterable, List, Optional

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, Request, UploadFile, status
from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

from ..repository import student_portal_repository as portal_repo
from ..repository import student_management_repository as roster_repo

from ..schemas import ResponseBase, WorkType
from ..schemas.student_portal_schema import StudentProfileCreate, StudentProfileResponse
from ..schemas.student_portal_schema import StudentProfileCreate, StudentProfileResponse
from ..utils.dependencies import member_required
from ..utils.file_handler import save_uploaded_file
from ..utils.file_handler import save_uploaded_file
from ..utils.student_portal_security import hash_password
from pydantic import ValidationError
from pydantic import ValidationError

router = APIRouter(prefix="/student-management", tags=["Student Management"])

TEMPLATE_HEADERS = [
    "Enrollment Number",
    "First Name",
    "Middle Name",
    "Last Name",
    "Std",
    "Div",
]

ENROLLMENT_MIN_LENGTH = 11
ENROLLMENT_MAX_LENGTH = 14
AUTO_PASSWORD_FALLBACK = "stud"

EXCEL_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}
CSV_MIME_TYPES = {"text/csv", "application/csv", "application/vnd.ms-excel"}


def _stream_excel(headers: List[str]) -> StreamingResponse:
    buffer = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Student Roster"
    sheet.append(headers)
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="student_roster_template.xlsx"'},
    )


@router.get("/template")
async def download_template(current_user: dict = Depends(member_required(WorkType.STUDENT))):
    del current_user
    return _stream_excel(TEMPLATE_HEADERS)


def _normalize_value(value: str | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _generate_auto_password(first_name: str, enrollment_number: str) -> str:
    normalized_name = re.sub(r"[^A-Za-z]", "", first_name).lower()
    prefix = normalized_name[:4] if normalized_name else AUTO_PASSWORD_FALLBACK
    enrollment_digits = re.sub(r"[^0-9]", "", enrollment_number)
    suffix_source = enrollment_digits or enrollment_number
    return f"{prefix}{suffix_source[-4:]}"


def _decode_csv_payload(raw_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "utf-16", "utf-16le", "utf-16be", "iso-8859-1"):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise HTTPException(status_code=400, detail="CSV file encoding is not supported. Please export as UTF-8.")


def _duplicate_row_payload(
    *,
    row_number: int | None,
    enrollment_number: str,
    first_name: str,
    middle_name: Optional[str],
    last_name: Optional[str],
    std: str,
    division: Optional[str],
    auto_password: Optional[str],
    reason: str,
) -> dict:
    return {
        # "row_number": row_number,
        "enrollment_number": enrollment_number,
        "first_name": first_name,
        "middle_name": middle_name,
        "last_name": last_name,
        "std": std,
        "division": division,
        "auto_password": auto_password,
        "reason": reason,
    }


def _build_duplicate_report(rows: List[dict]) -> dict:
    buffer = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Duplicate Students"
    headers = ["Row Number", *TEMPLATE_HEADERS, "Auto Password", "Reason"]
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                row.get("row_number"),
                row.get("enrollment_number"),
                row.get("first_name"),
                row.get("middle_name"),
                row.get("last_name"),
                row.get("std"),
                row.get("division"),
                row.get("auto_password"),
                row.get("reason"),
            ]
        )
    workbook.save(buffer)
    encoded = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return {"filename": "student_roster_duplicates.xlsx", "content": encoded}


@router.post("/upload", response_model=ResponseBase)
async def upload_student_roster(
    file: UploadFile = File(...),
    current_user: dict = Depends(member_required(WorkType.STUDENT)),
) -> ResponseBase:
    admin_id = current_user["admin_id"]
    filename = (file.filename or "").lower()
    content_type = (file.content_type or "").lower()
    raw_bytes = await file.read()
    if not raw_bytes:
        raise HTTPException(status_code=400, detail="File is empty")
    duplicate_rows: List[dict] = []
    def _record_duplicate(**kwargs) -> None:
        duplicate_rows.append(_duplicate_row_payload(**kwargs))
    def _process_rows(rows: Iterable[dict]) -> List[dict]:
        entries: List[dict] = []
        seen_enrollments: set[str] = set()
        for index, row in enumerate(rows, start=2):
            enrollment_number = _normalize_value(row.get("Enrollment Number"))
            first_name = _normalize_value(row.get("First Name"))
            middle_name = _normalize_value(row.get("Middle Name")) or None
            last_name = _normalize_value(row.get("Last Name")) or None
            std = _normalize_value(row.get("Std"))
            division = _normalize_value(row.get("Div")) or None
            if not enrollment_number:
                raise HTTPException(status_code=400, detail=f"Row {index}: Enrollment Number is required")
            if not (ENROLLMENT_MIN_LENGTH <= len(enrollment_number) <= ENROLLMENT_MAX_LENGTH):
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"Row {index}: Enrollment Number must be between "
                        f"{ENROLLMENT_MIN_LENGTH} and {ENROLLMENT_MAX_LENGTH} characters"
                    ),
                )
            if enrollment_number in seen_enrollments:
                _record_duplicate(
                    row_number=index,
                    enrollment_number=enrollment_number,
                    first_name=first_name,
                    middle_name=middle_name,
                    last_name=last_name,
                    std=std,
                    division=division,
                    auto_password=_generate_auto_password(first_name, enrollment_number),
                    reason="Duplicate enrollment number in file",
                )
                continue
            if not first_name or not std:
                raise HTTPException(status_code=400, detail=f"Row {index}: First Name and Std are required")
            seen_enrollments.add(enrollment_number)
            entries.append(
                {
                    "row_number": index,
                    "enrollment_number": enrollment_number,
                    "first_name": first_name,
                    "middle_name": middle_name,
                    "last_name": last_name,
                    "std": std,
                    "division": division,
                    "auto_password": _generate_auto_password(first_name, enrollment_number),
                }
            )
        return entries
    def _validate_headers(headers: List[str]) -> None:
        normalized_headers = [header.strip() for header in headers]
        if normalized_headers != TEMPLATE_HEADERS:
            raise HTTPException(status_code=400, detail="Headers must be: " + ", ".join(TEMPLATE_HEADERS))
    is_csv_like = content_type in CSV_MIME_TYPES or filename.endswith(".csv")
    is_excel_like = content_type in EXCEL_MIME_TYPES or filename.endswith(".xlsx")
    if is_csv_like:
        text = _decode_csv_payload(raw_bytes)
        if "\x00" in text:
            is_excel_like = True
        else:
            reader = csv.DictReader(io.StringIO(text))
            if not reader.fieldnames:
                raise HTTPException(status_code=400, detail="Missing CSV headers")
            _validate_headers(reader.fieldnames)
            entries = _process_rows(reader)
            is_excel_like = False
    if is_excel_like:
        try:
            workbook = load_workbook(filename=io.BytesIO(raw_bytes), data_only=True)
        except Exception as exc:  # pragma: no cover - openpyxl errors vary
            raise HTTPException(status_code=400, detail=f"Unable to read Excel file: {exc}") from exc
        sheet = workbook.active
        header_row = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        _validate_headers(header_row)
        excel_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(cell in (None, "") for cell in row):
                continue
            row_dict = {header: (row[idx] if idx < len(row) else None) for idx, header in enumerate(header_row)}
            excel_rows.append(row_dict)
        if not excel_rows:
            raise HTTPException(status_code=400, detail="No rows found in Excel")
        entries = _process_rows(excel_rows)
    elif not is_csv_like:
        raise HTTPException(status_code=400, detail="Only CSV or XLSX files are supported")
    if not entries:
        raise HTTPException(status_code=400, detail="No rows found in file")
    existing = roster_repo.fetch_existing_enrollments(
        admin_id,
        [entry["enrollment_number"] for entry in entries],
    )
    if existing:
        existing_set = set(existing)
        filtered_entries: List[dict] = []
        for entry in entries:
            if entry["enrollment_number"] in existing_set:
                _record_duplicate(
                    row_number=entry.get("row_number"),
                    enrollment_number=entry["enrollment_number"],
                    first_name=entry["first_name"],
                    middle_name=entry.get("middle_name"),
                    last_name=entry.get("last_name"),
                    std=entry["std"],
                    division=entry.get("division"),
                    auto_password=entry.get("auto_password"),
                    reason="Enrollment already exists",
                )
                continue
            filtered_entries.append(entry)
        entries = filtered_entries
    for entry in entries:
        entry.pop("row_number", None)
        entry["assigned_member_id"] = current_user["id"]
    if entries:
        roster_repo.insert_roster_entries(admin_id, entries)
        portal_repo.bulk_upsert_student_accounts(
            [
                {
                    "enrollment_number": entry["enrollment_number"],
                    "password_hash": hash_password(entry["auto_password"]),
                }
                for entry in entries
            ]
        )
    duplicate_report = _build_duplicate_report(duplicate_rows) if duplicate_rows else None
    if entries and duplicate_rows:
        message = "Uploaded with partial success. Some entries were duplicates."
    elif entries:
        message = "Student roster uploaded successfully"
    elif duplicate_rows:
        message = "No new students added because all records were duplicates."
    else:
        message = "No data processed."
    response_data: Dict[str, Any] = {
        "records_added": len(entries),
        "students": entries,
        "duplicate_count": len(duplicate_rows),
        "duplicates": duplicate_rows,
    }
    if duplicate_report:
        response_data["duplicate_report"] = duplicate_report
    return ResponseBase(status=bool(entries), message=message, data=response_data)


@router.post("/single", response_model=ResponseBase)
async def create_single_roster_student(
    payload: Dict[str, Any] = Body(...),
    current_user: dict = Depends(member_required(WorkType.STUDENT)),
) -> ResponseBase:
    """Create a single roster student from JSON row data.

    Expected keys (case-insensitive variants supported):
      - enrollment_number / Enrollment Number
      - first_name / First Name
      - last_name / Last Name
      - std / class / class_name / Std
      - division / Div
    """

    admin_id = current_user["admin_id"]

    # Support both snake_case and template-style keys
    enrollment_number = _normalize_value(
        payload.get("enrollment_number")
        or payload.get("Enrollment Number")
    )
    first_name = _normalize_value(
        payload.get("first_name")
        or payload.get("First Name")
    )

    # Middle name is optional; support both snake_case and template-style key
    middle_name_raw = (
        payload.get("middle_name")
        or payload.get("Middle Name")
    )
    middle_name = _normalize_value(middle_name_raw) or None
    last_name_raw = (
        payload.get("last_name")
        or payload.get("Last Name")
    )
    last_name = _normalize_value(last_name_raw) or None

    raw_std = (
        payload.get("std")
        or payload.get("class")
        or payload.get("class_name")
        or payload.get("Std")
    )
    std_val = _normalize_value(raw_std) if raw_std is not None else ""

    raw_div = payload.get("division") or payload.get("Div")
    division = _normalize_value(raw_div) or None

    # Validations similar to upload_student_roster
    if not enrollment_number:
        raise HTTPException(status_code=400, detail="Enrollment Number is required")

    if not (ENROLLMENT_MIN_LENGTH <= len(enrollment_number) <= ENROLLMENT_MAX_LENGTH):
        raise HTTPException(
            status_code=400,
            detail=(
                f"Enrollment Number must be between "
                f"{ENROLLMENT_MIN_LENGTH} and {ENROLLMENT_MAX_LENGTH} characters"
            ),
        )

    if not first_name or not std_val:
        raise HTTPException(status_code=400, detail="First Name and Std are required")

    # Duplicate check in DB
    existing = roster_repo.fetch_existing_enrollments(admin_id, [enrollment_number])
    if existing:
        raise HTTPException(status_code=400, detail="Enrollment already exists")

    auto_password = _generate_auto_password(first_name, enrollment_number)

    entry = {
        "enrollment_number": enrollment_number,
        "first_name": first_name,
        "middle_name": middle_name,
        "last_name": last_name,
        "std": std_val,
        "division": division,
        "auto_password": auto_password,
        "assigned_member_id": current_user["id"],
    }

    roster_repo.insert_roster_entries(admin_id, [entry])

    portal_repo.bulk_upsert_student_accounts(
        [
            {
                "enrollment_number": enrollment_number,
                "password_hash": hash_password(auto_password),
            }
        ]
    )

    response_data: Dict[str, Any] = {
        "records_added": 1,
        "students": [entry],
        "duplicate_count": 0,
        "duplicates": [],
    }

    return ResponseBase(status=True, message="Student added successfully", data=response_data)

@router.get("/roster", response_model=ResponseBase)
async def list_student_roster(current_user: dict = Depends(member_required(WorkType.STUDENT))) -> ResponseBase:
    admin_id = current_user["admin_id"]
    students = roster_repo.fetch_roster_entries(admin_id, member_id=current_user["id"])
    return ResponseBase(status=True, message="Student roster fetched successfully", data={"students": students})


@router.get("/roster/credentials", response_model=ResponseBase)
async def list_student_credentials(
    current_user: dict = Depends(member_required(WorkType.STUDENT)),
) -> ResponseBase:
    """Return enrollment numbers with their generated passwords only."""

    admin_id = current_user["admin_id"]
    roster_entries = roster_repo.fetch_roster_entries(admin_id, member_id=current_user["id"])

    credentials = [
        {
            "enrollment_number": entry.get("enrollment_number"),
            "password": entry.get("auto_password"),
        }
        for entry in roster_entries
    ]

    return ResponseBase(
        status=True,
        message="Student credentials fetched successfully",
        data={"students": credentials, "count": len(credentials)},
    )


@router.get("/roster/{enrollment_number}", response_model=ResponseBase)
async def get_roster_student(
    enrollment_number: str,
    current_user: dict = Depends(member_required(WorkType.STUDENT)),
) -> ResponseBase:
    normalized = enrollment_number.strip()
    if not normalized:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Enrollment number is required")

    admin_id = current_user["admin_id"]
    roster_entry = roster_repo.fetch_roster_entry_by_enrollment(
        admin_id,
        normalized,
        member_id=current_user["id"],
    )
    if not roster_entry:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Student not found")

    profile = portal_repo.get_student_profile_by_enrollment(normalized)
    return ResponseBase(
        status=True,
        message="Student fetched successfully",
        data={"profile": profile},
    )


@router.delete("/roster/{enrollment_number}", response_model=ResponseBase)
async def delete_roster_student(
    enrollment_number: str,
    current_user: dict = Depends(member_required(WorkType.STUDENT)),
) -> ResponseBase:
    normalized = enrollment_number.strip()
    if not normalized:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Enrollment number is required")

    admin_id = current_user["admin_id"]
    deleted = roster_repo.delete_roster_entry(
        admin_id,
        normalized,
        member_id=current_user["id"],
    )
    if not deleted:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Student not found in roster")

    return ResponseBase(status=True, message="Student removed from roster", data={"enrollment_number": normalized})


def _strip_optional(value: Optional[str]) -> Optional[str]:
# def _strip_optional(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    cleaned = value.strip()
    return cleaned or None


def _clean_filter(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    cleaned = value.strip()
    if not cleaned:
        return None
    if cleaned.lower() in {"none", "null", "undefined"}:
        return None
    return cleaned


def _is_upload_file(value: Any) -> bool:
    return hasattr(value, "filename") and hasattr(value, "content_type")


def _sanitize_upload_references(value: Any) -> Any:
    if _is_upload_file(value):
        return "<uploaded file>"
    if isinstance(value, dict):
        return {key: _sanitize_upload_references(val) for key, val in value.items()}
    if isinstance(value, list):
        return [_sanitize_upload_references(item) for item in value]
    return value


async def _parse_student_profile_payload(request: Request) -> StudentProfileCreate:
    content_type = (request.headers.get("content-type") or "").lower()

    try:
        # ----- JSON Payload -----
        if "application/json" in content_type:
            data = await request.json()

        # ----- FORM-DATA Payload -----
        elif "multipart/form-data" in content_type or "application/x-www-form-urlencoded" in content_type:
            form = await request.form()
            data: dict[str, object] = {}
            photo_upload: UploadFile | None = None

            for key, value in form.multi_items():
                # If this field is a file
                if isinstance(value, UploadFile):
                    if key in {"photo", "photo_path"}:
                        photo_upload = value
                    # NEVER pass UploadFile into Pydantic
                    continue

                # If field name is photo/photo_path (string version),
                # ignore to avoid conflict with file upload
                if key in {"photo", "photo_path"}:
                    continue

                data[key] = value

            # Save uploaded file (if any)
            if photo_upload:
                upload_info = await save_uploaded_file(photo_upload, subfolder="student-profiles")
                data["photo_path"] = upload_info["file_path"]

        # ----- Unsupported Content Type -----
        else:
            raise HTTPException(
                status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
                detail="Unsupported Content-Type. Use JSON or form-data",
            )

    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Malformed request body")
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Failed to parse request body")

    # ----- VALIDATE WITH Pydantic -----
    try:
        return StudentProfileCreate(**data)

    except ValidationError as exc:
        # Clean UploadFile from error output
        sanitized_errors = []
        for err in exc.errors():
            err_copy = dict(err)
            if isinstance(err_copy.get("input"), UploadFile):
                err_copy["input"] = "<uploaded file>"
            sanitized_errors.append(err_copy)

        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=sanitized_errors)


@router.put("/roster/{enrollment_number}", response_model=ResponseBase)
async def update_roster_student(
    enrollment_number: str,
    payload: StudentProfileCreate = Depends(_parse_student_profile_payload),
    current_user: dict = Depends(member_required(WorkType.STUDENT)),
) -> ResponseBase:

    # -------------------------
    # 1) Validate enrollment no
    # -------------------------
    normalized = enrollment_number.strip()
    if not normalized:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Enrollment number is required"
        )

    body_enrollment = payload.enrollment_number.strip()
    if body_enrollment and body_enrollment != normalized:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Enrollment number in body must match the URL parameter",
        )

    admin_id = current_user["admin_id"]

    # -------------------------
    # 2) Update roster fields
    # -------------------------
    roster_fields = {
        "first_name": payload.first_name.strip(),
        "last_name": _strip_optional(payload.father_name),
        "std": payload.class_stream.strip(),
        "division": _strip_optional(payload.division),
    }

    updated_roster = roster_repo.update_roster_entry(
        admin_id,
        normalized,
        **{key: value for key, value in roster_fields.items() if value is not None},
    )

    if not updated_roster:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student not found in roster"
        )

    # -------------------------
    # 3) Prepare profile payload
    # -------------------------
    profile_payload = payload.model_dump()
    profile_payload["enrollment_number"] = normalized

    # -------------------------
    # 4) Create or update profile
    # -------------------------
    existing_profile = portal_repo.get_student_profile_by_enrollment(normalized)

    if existing_profile:
        update_fields = {
            key: value
            for key, value in profile_payload.items()
            if key != "enrollment_number" and value is not None
        }

        if update_fields:
            profile_record = portal_repo.update_student_profile(
                existing_profile["id"], **update_fields
            )
            profile_data = profile_record or existing_profile
        else:
            profile_data = existing_profile
    else:
        # Do not auto-create a new profile when updating roster.
        # Simply keep profile data as None or existing_profile (which is already None).
        profile_data = existing_profile

    # -------------------------
    # 5) FINAL RESPONSE (ONLY PROFILE)
    # -------------------------
    return ResponseBase(
        status=True,
        message="Student updated successfully",
        data=profile_data,
    )


def _clean_filter(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    cleaned = value.strip()
    if not cleaned:
        return None
    if cleaned.lower() in {"none", "null", "undefined"}:
        return None
    return cleaned


def _is_upload_file(value: Any) -> bool:
    return hasattr(value, "filename") and hasattr(value, "content_type")


def _sanitize_upload_references(value: Any) -> Any:
    if _is_upload_file(value):
        return "<uploaded file>"
    if isinstance(value, dict):
        return {key: _sanitize_upload_references(val) for key, val in value.items()}
    if isinstance(value, list):
        return [_sanitize_upload_references(item) for item in value]
    return value


async def _parse_student_profile_payload(request: Request) -> StudentProfileCreate:
    print(">>> CONTENT TYPE:", request.headers.get("content-type"))

    content_type = request.headers.get("content-type", "").lower()

    if "multipart/form-data" in content_type:
        form = await request.form()
        data = {}
        photo_file = None

        # MOST IMPORTANT
        for key, value in form.multi_items():

            if isinstance(value, UploadFile):
                if key in ("photo", "photo_path"):
                    photo_file = value
                continue

            if key in ("photo", "photo_path"):
                continue

            data[key] = value

        # Save the uploaded image
        if photo_file and photo_file.filename.strip():
            print(">>> FILE RECEIVED:", photo_file.filename)
            saved = await save_uploaded_file(photo_file, "student-profiles")
            data["photo_path"] = saved["file_path"]
        else:
            data["photo_path"] = None


        return StudentProfileCreate(**data)

    elif "application/json" in content_type:
        body = await request.json()
        return StudentProfileCreate(**body)

    else:
        raise HTTPException(
            415,
            "Unsupported content type. Send JSON or form-data",
        )


@router.put("/roster/{enrollment_number}", response_model=ResponseBase)
async def update_roster_student(
    enrollment_number: str,
    payload: StudentProfileCreate = Depends(_parse_student_profile_payload),
    current_user: dict = Depends(member_required(WorkType.STUDENT)),
) -> ResponseBase:

    # -------------------------
    # 1) Validate enrollment no
    # -------------------------
    normalized = enrollment_number.strip()
    if not normalized:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Enrollment number is required"
        )

    body_enrollment = payload.enrollment_number.strip()
    if body_enrollment and body_enrollment != normalized:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Enrollment number in body must match the URL parameter",
        )

    admin_id = current_user["admin_id"]

    # -------------------------
    # 2) Update roster fields
    # -------------------------
    roster_fields = {
        "first_name": payload.first_name.strip(),
        "last_name": _strip_optional(payload.father_name),
        "std": payload.class_stream.strip(),
        "division": _strip_optional(payload.division),
    }

    updated_roster = roster_repo.update_roster_entry(
        admin_id,
        normalized,
        **{key: value for key, value in roster_fields.items() if value is not None},
    )

    if not updated_roster:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student not found in roster"
        )

    # -------------------------
    # 3) Prepare profile payload
    # -------------------------
    profile_payload = payload.model_dump()
    profile_payload["enrollment_number"] = normalized

    # Ensure empty string is not stored
    if payload.photo_path == "" or payload.photo_path is None:
        profile_payload.pop("photo_path", None)


    # -------------------------
    # 4) Create or update profile
    # -------------------------
    existing_profile = portal_repo.get_student_profile_by_enrollment(normalized)

    if existing_profile:
        update_fields = {
            key: value
            for key, value in profile_payload.items()
            if key != "enrollment_number" and value is not None
        }

        if update_fields:
            profile_record = portal_repo.update_student_profile(
                existing_profile["id"], **update_fields
            )
            profile_data = profile_record or existing_profile
        else:
            profile_data = existing_profile
    else:
        profile_data = portal_repo.create_student_profile(**profile_payload)

    # -------------------------
    # 5) FINAL RESPONSE (ONLY PROFILE)
    # -------------------------
    return ResponseBase(
        status=True,
        message="Student updated successfully",
        data=profile_data,
    )


@router.get("/students", response_model=ResponseBase)
async def list_student_profiles(
    class_filter: Optional[str] = Query(default=None, alias="class_filter"),
    division_filter: Optional[str] = Query(default=None, alias="division_filter"),
    class_alias: Optional[str] = Query(default=None, alias="class"),
    division_alias: Optional[str] = Query(default=None, alias="division"),
    current_user: dict = Depends(member_required(WorkType.STUDENT)),
) -> ResponseBase:
    admin_id = current_user["admin_id"]
    resolved_class = _clean_filter(class_filter) or _clean_filter(class_alias)
    resolved_division = _clean_filter(division_filter) or _clean_filter(division_alias)
    students = roster_repo.fetch_student_profiles(
        admin_id,
        class_filter=resolved_class,
        division_filter=resolved_division,
        member_id=current_user["id"]
    )
    return ResponseBase(status=True, message="Student profiles fetched successfully", data={"students": students})


@router.get("/students/filters", response_model=ResponseBase)
async def list_student_filters(
    current_user: dict = Depends(member_required(WorkType.STUDENT)),
) -> ResponseBase:
    admin_id = current_user["admin_id"]
    filters = roster_repo.fetch_class_division_filters(admin_id, member_id=current_user["id"])
    return ResponseBase(
        status=True,
        message="Class and division filters fetched successfully",
        data=filters,
    )
@router.get("/subjects", response_model=ResponseBase)
async def list_class_subjects(
    class_filter: Optional[str] = Query(default=None, alias="class_filter"),
    division_filter: Optional[str] = Query(default=None, alias="division_filter"),
    class_alias: Optional[str] = Query(default=None, alias="class"),
    division_alias: Optional[str] = Query(default=None, alias="division"),
    current_user: dict = Depends(member_required(WorkType.STUDENT)),
) -> ResponseBase:
    admin_id = current_user["admin_id"]
    resolved_class = _clean_filter(class_filter) or _clean_filter(class_alias)
    if not resolved_class:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Class parameter is required")
    resolved_division = _clean_filter(division_filter) or _clean_filter(division_alias)
    subjects = roster_repo.fetch_class_subjects(
        admin_id,
        class_filter=resolved_class,
        division_filter=resolved_division,
        member_id=current_user["id"]
    )
    return ResponseBase(status=True, message="Subjects fetched successfully", data={"subjects": subjects})


@router.get("/subjects", response_model=ResponseBase)
async def list_class_subjects(
    class_filter: Optional[str] = Query(default=None, alias="class_filter"),
    division_filter: Optional[str] = Query(default=None, alias="division_filter"),
    class_alias: Optional[str] = Query(default=None, alias="class"),
    division_alias: Optional[str] = Query(default=None, alias="division"),
    current_user: dict = Depends(member_required(WorkType.STUDENT)),
) -> ResponseBase:
    admin_id = current_user["admin_id"]
    resolved_class = _clean_filter(class_filter) or _clean_filter(class_alias)
    if not resolved_class:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Class parameter is required")
    resolved_division = _clean_filter(division_filter) or _clean_filter(division_alias)
    subjects = roster_repo.fetch_class_subjects(
        admin_id,
        class_filter=resolved_class,
        division_filter=resolved_division,
        member_id=current_user["id"]
    )
    return ResponseBase(status=True, message="Subjects fetched successfully", data={"subjects": subjects})